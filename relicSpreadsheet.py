import pandas as pd
import requests
from bs4 import BeautifulSoup
from lxml import html
from openpyxl import load_workbook

def fetch_and_parse(url, xpath_expression):
    # Fetch the HTML content
    response = requests.get(url)
    html_content = response.content
    
    # Parse the HTML content
    tree = html.fromstring(html_content)
    
    # Find elements using XPath
    elements = tree.xpath(xpath_expression)
    
    return [element.text_content() for element in elements]

def webScrape(excelList, baseurl, ws):
    excount = 2
    for cell in excelList:
        if type(cell[0]) == str:
            url = baseurl + cell[0] + "_" + cell[1]
            
            response = requests.get(url)
            html_content = response.content

            tree = html.fromstring(html_content)

            print(excount)
            
            for i in range(6):
                xpath = ('//*[@id="72656C6963table"]/tbody/tr[' + str(2+i) + ']/td[1]/a[2]')
                texts = fetch_and_parse(url, xpath)
                
                for text in texts:
                    text = text
                if i == 0:
                    ws['D' + str(excount)] = text
                elif i == 1:
                    ws['F' + str(excount)] = text
                elif i == 2:
                    ws['H' + str(excount)] = text
                elif i == 3:
                    ws['J' + str(excount)] = text
                elif i == 4:
                    ws['L' + str(excount)] = text
                elif i == 5:
                    ws['M' + str(excount)] = text
            excount+=1
            
    
def main():
    wb = load_workbook('warframeRelics.xlsx')
    ws = wb.active

    
    exFile = 'warframeRelics.xlsx'

    baseurl = 'https://warframe.fandom.com/wiki/'

    #Pandas collects data as a DataFrame
    df_columns = pd.read_excel(exFile, usecols='A:B')

    #.tolist() converts DataFrame to list of lists
    excelList = df_columns.values.tolist()

    webScrape(excelList, baseurl, ws)

    wb.save('warframeRelics.xlsx')
    print("saved")

if __name__=="__main__":
    main()
