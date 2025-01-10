import pandas as pd
import requests
from bs4 import BeautifulSoup
from lxml import html
from openpyxl import load_workbook
from pathlib import Path
import json
import os
import sys

def startup():
    check = Path("relicData.json")
    if check.exists():
        print("Checking for updates.")
        checkForUpdate()
        os.system("start EXCEL.EXE warframeRelics.xlsx")
    else:
        print("Save file not found.")
        print("Creating save file.")
        checkForUpdate()
        os.system("start EXCEL.EXE warframeRelics.xlsx")
    #if there is run checkUpdate()
    #if not skip straight to scrapeForRelics and save to json


def fetch_and_parse(url, xpath_expression):
    # Fetch the HTML content
    response = requests.get(url)
    html_content = response.content
    
    # Parse the HTML content
    tree = html.fromstring(html_content)
    
    # Find elements using XPath
    elements = tree.xpath(xpath_expression)
    
    return [element.text_content() for element in elements]

def scrapeForRelics():
    url = "https://warframe.fandom.com/wiki/Void_Relic"
    unvaultedXPath = '//*[@id="mw-content-text"]/div[1]/table[2]/tbody/tr[2]'
    vaultedXPath = '//*[@id="mw-customcollapsible-VaultedRelicList"]/table/tbody/tr[2]'

    relics = {}

    unvaultedTexts = listify(fetch_and_parse(url, unvaultedXPath))
    for relic in unvaultedTexts:
            relics[relic] = False
    vaultedTexts = listify(fetch_and_parse(url, vaultedXPath))
    for relic in vaultedTexts:
            relics[relic] = True
        
       
    for text in vaultedTexts:
        #replace \xao (space) character with space
        text = text.replace(u'\xa0', u' ')
        #split to list when new line
        text = text.splitlines()
        #remove empty list elements
        text = list(filter(None, text))
        for relic in text:
            relics[relic] = True
    
    return relics

def saveRelicData(relics):
    with open("relicData.json", "w") as outfile:
        json.dump(relics, outfile)

def checkForUpdate():
    #(might be better to scrape data from the spreadsheet instead or get the list of relics and compair them to the excel)
    with open('relicData.json', 'r') as infile:
        dataset1 = json.load(infile)
    relics = scrapeForRelics()

    if dataset1 == relics:
        print("Save file up to date.")
    else:
        print("Save file out of date.")
        print("Updating...")
        try:
            loadRelicData(relics)
            print()
            print("Data saved.")
        except:
            print()
            print("An error has occured. Please try again.")
            sys.exit()
        relics = scrapeForRelics()
        saveRelicData(relics)

def listify(texts):
    for text in texts:
        #replace \xao (space) character with space
        text = text.replace(u'\xa0', u' ')
        #split to list when new line
        text = text.splitlines()
        #remove empty list elements
        text = list(filter(None, text))
    return text

def fetch_and_parse(url, xpath_expression):
    # Fetch the HTML content
    response = requests.get(url)
    html_content = response.content
    
    # Parse the HTML content
    tree = html.fromstring(html_content)
    
    # Find elements using XPath
    elements = tree.xpath(xpath_expression)
    
    return [element.text_content() for element in elements]

def webScrape(excelList, baseurl, ws, relics):
    excount = 2
    percentageCount = 0
    for cell in excelList:
        if type(cell[0]) == str:
            url = baseurl + cell[0] + "_" + cell[1]
            
            response = requests.get(url)
            html_content = response.content

            tree = html.fromstring(html_content)

            print(f"\rProgress: {round(percentageCount/len(excelList)*100)}%", end="")
            
            for i in range(6):
                xpath = ('//*[@id="72656C6963table"]/tbody/tr[' + str(2+i) + ']/td[1]/a[2]')
                texts = fetch_and_parse(url, xpath)
                
                for text in texts:
                    text = text
                if i == 0:
                    ws['D' + str(excount)] = text
                elif i == 1:
                    ws['F' + str(excount)] = text
                elif i == 2:
                    ws['H' + str(excount)] = text
                elif i == 3:
                    ws['J' + str(excount)] = text
                elif i == 4:
                    ws['L' + str(excount)] = text
                elif i == 5:
                    ws['M' + str(excount)] = text
            excount+=1
            percentageCount+=1
    
def loadRelicData(relics):
    wb = load_workbook('warframeRelics.xlsx')
    ws = wb.active

    exFile = 'warframeRelics.xlsx'

    baseurl = 'https://warframe.fandom.com/wiki/'

    #Pandas collects data as a DataFrame
    df_columns = pd.read_excel(exFile, usecols='A:B')

    #.tolist() converts DataFrame to list of lists
    excelList = df_columns.values.tolist()

    webScrape(excelList, baseurl, ws, relics)

    wb.save('warframeRelics.xlsx')

if __name__=="__main__":
    startup()
